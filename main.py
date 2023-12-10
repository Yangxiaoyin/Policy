import openpyxl as vb
from openpyxl import Workbook, load_workbook
from datetime import datetime, timedelta
import xlsxwriter
import math
import random
from scipy.stats import norm


# Read assumptionsm and assign them as input Data object
class inputData:
    def __init__(self, Product, Step_Up, Step_Up_Period, Rider_Charge, Initial_Premium, First_Withdrawal_Age,
                 Annuity_Commencement_Date,
                 Last_Death_Age, Mortality, Withdrawal_Rate, Fixed_Allocation_Funds_Automatic_Rebalancing_Target, MandE,
                 Fund_Fees,
                 Risk_Free_Rate, Volatility, MAW_Age1, MAW_Age2, MAW_Age3, MAW_Age4, MAW_Rate1, MAW_Rate2, MAW_Rate3,
                 MAW_Rate4):
        self.Product = Product
        self.Step_Up = Step_Up
        self.Step_Up_Period = Step_Up_Period
        self.Rider_Charge = Rider_Charge
        self.Initial_Premium = Initial_Premium
        self.First_Withdrawal_Age = First_Withdrawal_Age
        self.Annuity_Commencement_Date = Annuity_Commencement_Date
        self.Last_Death_Age = Last_Death_Age
        self.Mortality = Mortality
        self.Withdrawal_Rate = Withdrawal_Rate
        self.Fixed_Allocation_Funds_Automatic_Rebalancing_Target = Fixed_Allocation_Funds_Automatic_Rebalancing_Target
        self.MandE = MandE
        self.Fund_Fees = Fund_Fees
        self.Risk_Free_Rate = Risk_Free_Rate
        self.Volatility = Volatility
        self.MAW_Age1 = MAW_Age1
        self.MAW_Age2 = MAW_Age2
        self.MAW_Age3 = MAW_Age3
        self.MAW_Age4 = MAW_Age4
        self.MAW_Rate1 = MAW_Rate1
        self.MAW_Rate2 = MAW_Rate2
        self.MAW_Rate3 = MAW_Rate3
        self.MAW_Rate4 = MAW_Rate4

    # 显示所有数据
    def printAllInputData(self):
        print(self.Product)
        print(self.Step_Up)
        print(self.Step_Up_Period)
        print(self.Rider_Charge)
        print(self.Initial_Premium)
        print(self.First_Withdrawal_Age)
        print(self.Annuity_Commencement_Date)
        print(self.Last_Death_Age)
        print(self.Mortality)
        print(self.Withdrawal_Rate)
        print(self.Fixed_Allocation_Funds_Automatic_Rebalancing_Target)
        print(self.MandE)
        print(self.Fund_Fees)
        print(self.Risk_Free_Rate)
        print(self.Volatility)
        print(self.MAW_Age1)
        print(self.MAW_Age2)
        print(self.MAW_Age3)
        print(self.MAW_Age4)
        print(self.MAW_Rate1)
        print(self.MAW_Rate2)
        print(self.MAW_Rate3)
        print(self.MAW_Rate4)


# create object policy as output in excel
class policy:
    # A
    Year = []
    # B
    Anniversary = []
    # C
    Age = []
    # D
    Contribution = []
    # E
    AV_Pre_Fee = []
    # F
    Fund1_Pre_Fee = []
    # G
    Fund2_Pre_Fee = []
    # H
    MandE_Fund_Fees = []
    # I
    AV_Pre_Withdrawal = []
    # J
    Fund1_Pre_Withdrawal = []
    # K
    Fund2_Pre_Withdrawal = []
    # L and AD are same varaibles

    # Withdrawal_Amount = []
    # M
    AV_Post_Withdrawal = []
    # N
    Fund1_Post_Withdrawal = []
    # O
    Fund2_Post_Withdrawal = []
    # P
    Rider_Charge = []
    # Q
    AV_Post_Charges = []
    # R
    Fund1_Post_Charges = []
    # S
    Fund2_Post_Charges = []
    # T
    Death_Payment = []
    # U
    AV_Post_Death_Claims = []
    # V
    Fund1_Post_Death_Claims = []
    # W
    Fund2_Post_Death_Claims = []
    # X
    Fund1_Post_Rebalance = []
    # Y
    Fund2_Post_Rebalance = []
    # Z
    ROP_Death_Base = []

    # AA
    NAR_Death_Claims = []
    # AB
    Death_Benefit_base = []
    # AC
    Withdrawal_Base = []
    # AD
    Withdrawal_Amount = []
    # AE
    Cumulative_Withdrawal = []
    # AF
    Maximum_Annual_Withdrawal = []
    # AG
    Maximum_Annual_Withdrawal_Rate = []

    # AI
    Eligible_Step_Up = []
    # AJ
    Growth_Phase = []
    # AK
    Withdrawal_Phase = []
    # AL
    Automatic_Periodic_Benefit_Status = []
    # AM
    Last_Death = []

    # AO
    Fund1_Return = []
    # AP
    Fund2_Return = []
    # AQ
    Rebalance_Indicator = []
    # AR
    DF = []

    # AT
    qx = []
    # AU
    Death_Claims = []
    # AV
    Withdrawal_Claims = []
    # AW
    Rider_Charges = []

    # AU16 PV_DB_Claim: DF*outData.Death_Claims
    PV_DB_Claim = 0
    # AV16 PV_WB_Claim: DF*outData.Death_Claims
    PV_WB_Claim = 0
    # AW16 PV_RC: DF*outData.Rider_Charges
    PV_RC = 0

    def outputToExcel(self):
        pass


wb = vb.load_workbook('Test VA Withdrawal.xlsx',data_only=True)
ws = wb["Main"]
wsMortalityRate = wb["DecrermentAssumptions"]

# Read from assumptions and store the data into inData
inData = inputData(ws['B2'].value, ws['B3'].value, ws['B4'].value, ws['B5'].value, ws['B7'].value, ws['B8'].value,
                   ws['B9'].value,
                   ws['B10'].value, ws['B11'].value, ws['B12'].value, ws['B13'].value, ws['E3'].value, ws['E4'].value,
                   ws['E6'].value,
                   ws['E7'].value, ws['G4'].value, ws['G5'].value, ws['G6'].value, ws['G7'].value, ws['H4'].value,
                   ws['H5'].value,
                   ws['H6'].value, ws['H7'].value)
#inData.printAllInputData()

# Define output Object
outData = policy()
# ==================Part ONE: to calculate C,D,Age,contribution; AJ,AI,AM(status table),AO,AP(Fund Return),AR(DF),AT(Mortality Rate)========
# Hard coded part: start year and max_index
max_index = 41
# Assign value to A(index)
outData.Year = list(range(max_index))
# Assign B as year
start_date = datetime(2016, 8, 1)
end_date = datetime(start_date.year + max_index - 1, start_date.month, start_date.day)
current_date = start_date
while current_date <= end_date:
    outData.Anniversary.append(current_date.strftime("%Y/%m/%d"))
    current_date = current_date.replace(year=current_date.year + 1)
# Assign C as age
outData.Age = list(range(60, 60 + max_index))
# Assign D as contribution (All assumed as 0 in the table)
for item in range(0, max_index):
    outData.Contribution.append(0)

# Assign AJ as Growth_phase
# =IF(AND(C19<=Age.FirstWD,C19<=Age.AnnuityComm,C19<Age.Death),1,0)
outData.Growth_Phase.append('')
for item in range(1, max_index):
    if outData.Age[item] <= inData.First_Withdrawal_Age and outData.Age[item] <= inData.Annuity_Commencement_Date and \
            outData.Age[item] <= inData.Last_Death_Age:
        outData.Growth_Phase.append(1)
    else:
        outData.Growth_Phase.append(0)

# Assign AI as step_up status,need AJ19
# AI=IF(AND(A19<=StepUp.Yr,AJ19=1),1,0)
outData.Eligible_Step_Up.append('')
for item in range(1, max_index):
    if outData.Year[item] <= inData.Step_Up_Period and outData.Growth_Phase[item] == 1:
        outData.Eligible_Step_Up.append(1)
    else:
        outData.Eligible_Step_Up.append(0)

# Assign AM as last death
outData.Last_Death.append('')
for item in range(1, max_index):
    if outData.Age[item] == inData.Last_Death_Age:
        outData.Last_Death.append(1)
    else:
        outData.Last_Death.append(0)

# Assign AO as fixed fund return
outData.Fund1_Return.append('')
for item in range(1, max_index):  # item starting with 1
    outData.Fund1_Return.append(inData.Risk_Free_Rate)

# Assign AP with normal distribution as Fund_Return2
outData.Fund2_Return.append('')
for item in range(1, max_index):
    random_number = random.random()  # 生成0到1之间的随机数
    norm_inverse = norm.ppf(random_number, loc=0, scale=1)  # 生成符合正态分布的随机数
    # =EXP(LN(1 +$E$6)-0.5 *$E$7 ^ 2 +$E$7 * NORMINV(RAND(), 0, 1))-1
    outData.Fund2_Return.append(math.exp(
        math.log(1 + inData.Risk_Free_Rate) - 0.5 * pow(inData.Volatility, 2) + inData.Volatility * norm_inverse) - 1)

# Assign AR as DiscountFactor
for item in range(0, max_index):
    outData.DF.append((1 + inData.Risk_Free_Rate) ** -outData.Year[item])


# Assign Value to AT, read qx from reference_assumption table
titles = []
for item in list(wsMortalityRate.rows)[1][0:2]:
# Read from excel as rows and convert it to lists,take age as key in dictionary
    titles.append(item.value)
    #print(item.value)

data_lists = []  # Define a total list to store data
for item in list(wsMortalityRate.rows)[2:]:  #Read excel from row3
    value_dict = {}  # put two dictionary into each list item
    value_dict[titles[0]] = item[0].value
    #value_dict[titles[1]] = inData.Mortality
    value_dict[titles[1]] = item[1].value
    data_lists.append(value_dict)  # store dict into the list
#print(data_lists)

for item in range(0, max_index):
    exit_flag = 0
    for dictionary in data_lists:
        if outData.Age[item] in dictionary.values():
            exit_flag = 1
            outData.qx.append(dictionary[list(dictionary.keys())[1]]) # append mortality to AT(qx)
            break
        if exit_flag == 1:
            outData.qx.append("FALSE")

# print(outData.qx)
# print(outData.Year)
# print(outData.Anniversary)
# print(outData.Age)
# print(outData.Contribution)
# ==========================================Part TWO: Initialization of first row =====================================

# F18
outData.Fund1_Pre_Fee.append(inData.Initial_Premium * 0.16)

# G18
outData.Fund2_Pre_Fee.append(inData.Initial_Premium * 0.64)

# E18 = F18+G18
outData.AV_Pre_Fee.append(outData.Fund1_Pre_Fee[0] + outData.Fund2_Pre_Fee[0])

# H18 = 0
outData.MandE_Fund_Fees.append(0)

# I18 = E18+D18-H18
outData.AV_Pre_Withdrawal.append(outData.AV_Pre_Fee[0] + outData.Contribution[0] - outData.MandE_Fund_Fees[0])

# J18=IF($I18=0,0,F18*($I18/$E18))
if outData.AV_Pre_Withdrawal[0] == 0:
    outData.Fund1_Pre_Withdrawal.append(0)
else:
    outData.Fund1_Pre_Withdrawal.append(
        outData.Fund1_Pre_Fee[0] * (outData.AV_Pre_Withdrawal[0] / outData.AV_Pre_Fee[0]))

# K18=IF($I18=0,0,G18*($I18/$E18))
if outData.AV_Pre_Withdrawal[0] == 0:
    outData.Fund2_Pre_Withdrawal.append(0)
else:
    outData.Fund2_Pre_Withdrawal.append(
        outData.Fund2_Pre_Fee[0] * (outData.AV_Pre_Withdrawal[0] / outData.AV_Pre_Fee[0]))

# M18 = I18
outData.AV_Post_Withdrawal.append(outData.AV_Pre_Withdrawal[0])

# N18=IF($M18=0,0,J18*($M18/$I18))
if outData.AV_Post_Withdrawal[0] == 0:
    outData.Fund1_Post_Withdrawal.append(0)
else:
    outData.Fund1_Post_Withdrawal.append(
        outData.Fund1_Pre_Withdrawal[0] * (outData.AV_Post_Withdrawal[0] / outData.AV_Pre_Withdrawal[0]))

# O18=IF($M18=0,0,K18*($M18/$I18))
if outData.AV_Post_Withdrawal[0] == 0:
    outData.Fund2_Post_Withdrawal.append(0)
else:
    outData.Fund2_Post_Withdrawal.append(
        outData.Fund2_Pre_Withdrawal[0] * (outData.AV_Post_Withdrawal[0] / outData.AV_Pre_Withdrawal[0]))

# P18
outData.Rider_Charge.append(0)

# Q18 = M18-P18
outData.AV_Post_Charges.append(outData.AV_Post_Withdrawal[0] - outData.Rider_Charge[0])

# R18=IF($Q18=0,0,N18*($Q18/$M18))
if outData.AV_Post_Charges[0] == 0:
    outData.Fund1_Post_Charges.append(0)
else:
    outData.Fund1_Post_Charges.append(
        outData.Fund1_Post_Withdrawal[0] * (outData.AV_Post_Charges[0] / outData.AV_Post_Withdrawal[0]))

# s18=IF($Q18=0,0,O18*($Q18/$M18))
if outData.AV_Post_Charges[0] == 0:
    outData.Fund2_Post_Charges.append(0)
else:
    outData.Fund2_Post_Charges.append(
        outData.Fund2_Post_Withdrawal[0] * (outData.AV_Post_Charges[0] / outData.AV_Post_Withdrawal[0]))

# T18
outData.Death_Payment.append(0)

# U18=MAX(Q18-T18,0)
if outData.AV_Post_Charges[0] - outData.Death_Payment[0] >= 0:
    outData.AV_Post_Death_Claims.append(outData.AV_Post_Charges[0] - outData.Death_Payment[0])
else:
    outData.AV_Post_Death_Claims.append(0)

# V18=IF($U18=0,0,R18*($U18/$Q18))
if outData.AV_Post_Death_Claims == 0:
    outData.Fund1_Post_Death_Claims.append(0)
else:
    outData.Fund1_Post_Death_Claims.append(
        outData.Fund1_Post_Charges[0] * (outData.AV_Post_Death_Claims[0] / outData.AV_Post_Charges[0]))

# W18==IF($U18=0,0,S18*($U18/$Q18))
if outData.AV_Post_Death_Claims == 0:
    outData.Fund2_Post_Death_Claims.append(0)
else:
    outData.Fund2_Post_Death_Claims.append(
        outData.Fund2_Post_Charges[0] * (outData.AV_Post_Death_Claims[0] / outData.AV_Post_Charges[0]))

# Z18
outData.ROP_Death_Base.append(inData.Initial_Premium)

# AA18=MAX(0,T18-Q18)
outData.NAR_Death_Claims.append(max(0, outData.Death_Payment[0] - outData.AV_Post_Charges[0]))

# AB18 = B7
outData.Death_Benefit_base.append(inData.Initial_Premium)

# AC18 = B7
outData.Withdrawal_Base.append(inData.Initial_Premium)

# AD18
outData.Withdrawal_Amount.append(0)

# L18 = AD18

# AE18=SUM(AD$18:AD18)
outData.Cumulative_Withdrawal.append(sum(outData.Withdrawal_Amount[:0]))

# AF18
outData.Maximum_Annual_Withdrawal.append(0)

# AG18
outData.Maximum_Annual_Withdrawal_Rate.append(0)

# AK18, need U
outData.Withdrawal_Phase.append('')

# AL18
outData.Automatic_Periodic_Benefit_Status.append('')

# AQ18 , we need AK, AL
outData.Rebalance_Indicator.append('')

# X18=IF(AQ18=1,U18*Fund.Reb.Target,V18)
if outData.Rebalance_Indicator[0] == 1:
    outData.Fund1_Post_Rebalance.append(
        outData.AV_Post_Death_Claims[0] * inData.Fixed_Allocation_Funds_Automatic_Rebalancing_Target)
else:
    outData.Fund1_Post_Rebalance.append(outData.Fund1_Post_Death_Claims[0])

# Y18=Q18-X18
outData.Fund2_Post_Rebalance.append(outData.AV_Post_Charges[0] - outData.Fund1_Post_Rebalance[0])

# AU18, AU=AA
outData.Death_Claims = outData.NAR_Death_Claims

# AV18
outData.Withdrawal_Claims.append(0)

# AW18 AW=P
outData.Rider_Charges = outData.Rider_Charge

# print(outData.Eligible_Step_Up)
# print(outData.Growth_Phase)
# print(outData.Withdrawal_Phase)
# print(outData.Last_Death)
# print(outData.Automatic_Periodic_Benefit_Status)
# ========================================== Part THREE: Build an Iterator==========================================================
# =====================================================第三部分==========================================================
# F19=X18*(1+AO19)
outData.Fund1_Pre_Fee.append(outData.Fund1_Post_Rebalance[0] * (1 + outData.Fund1_Return[1]))
# G19=Y18*(1+AP19)
outData.Fund2_Pre_Fee.append(outData.Fund2_Post_Rebalance[0] * (1 + outData.Fund2_Return[1]))
# E19=F19+G19
outData.AV_Pre_Fee.append(outData.Fund1_Pre_Fee[1] + outData.Fund2_Pre_Fee[1])
# H19=U18*(Rate.MandE+Rate.FundFee)
outData.MandE_Fund_Fees.append(outData.AV_Post_Death_Claims[0] * (inData.MandE + inData.Fund_Fees))
# I19=MAX(0,E19+D19-H19)
outData.AV_Pre_Withdrawal.append(max(0, outData.AV_Pre_Fee[1] + outData.Contribution[1] - outData.MandE_Fund_Fees[1]))

# AK19=IF(AND(OR(C19>Age.FirstWD,C19>Age.AnnuityComm),U18>0,C19<Age.Death),1,0)
if (outData.Age[1] > inData.First_Withdrawal_Age or outData.Age[1] > inData.Annuity_Commencement_Date) and \
        outData.AV_Post_Death_Claims[0] > 0 and outData.Age[1] < inData.Last_Death_Age:
    outData.Withdrawal_Phase.append(1)
else:
    outData.Withdrawal_Phase.append(0)

# AL19
outData.Automatic_Periodic_Benefit_Status.append(0)

# AG19=IF(AJ19=1,0,IF(C19>MAW.Age4,MAW.Rate4,IF(C19>MAW.Age3,MAW.Rate3,IF(C19>MAW.Age2,MAW.Rate2,IF(C19>MAW.Age1,MAW.Rate1,0)))))
if outData.Growth_Phase[1] == 1:
    outData.Maximum_Annual_Withdrawal_Rate.append(0)
elif outData.Age[1] > inData.MAW_Age4:
    outData.Maximum_Annual_Withdrawal_Rate.append(inData.MAW_Rate4)
elif outData.Age[1] > inData.MAW_Age3:
    outData.Maximum_Annual_Withdrawal_Rate.append(inData.MAW_Rate3)
elif outData.Age[1] > inData.MAW_Age2:
    outData.Maximum_Annual_Withdrawal_Rate.append(inData.MAW_Rate2)
elif outData.Age[1] > inData.MAW_Age1:
    outData.Maximum_Annual_Withdrawal_Rate.append(inData.MAW_Rate1)
else:
    outData.Maximum_Annual_Withdrawal_Rate.append(0)

# T19=IF(SUM(AJ19:AM19)=0,0,MAX(AB18,Z18)*AT19)
if outData.Growth_Phase[1] + outData.Withdrawal_Phase[1] + outData.Automatic_Periodic_Benefit_Status[1] + \
        outData.Last_Death[1] == 0:
    outData.Death_Payment.append(0)
else:
    outData.Death_Payment.append(max(outData.Death_Benefit_base[0], outData.ROP_Death_Base[0]) * outData.qx[1])

# ============================for calculating U,P, AC, we need an iterator =====

# To assign U19
outData.AV_Post_Death_Claims.append(outData.AV_Post_Death_Claims[0])

# AC19=MAX(IF(AJ19=1,U19,0),AC18*(1-AT19)+D19,IF(AI19=1,AC18*(1-AT19)*(1+Rate.StepUp)+D19-H19-P19,0))
outData.Withdrawal_Base.append(outData.Withdrawal_Base[0])

#  AF19=AG19*AC19
outData.Maximum_Annual_Withdrawal.append(outData.Maximum_Annual_Withdrawal_Rate[1] * outData.Withdrawal_Base[1])

#  AD19=IF(AK19=1,Rate.WD*AC19,IF(AL19=1,AF19,0))
if outData.Withdrawal_Phase[1] == 1:
    outData.Withdrawal_Amount.append(inData.Withdrawal_Rate * outData.Withdrawal_Base[1])
elif outData.Automatic_Periodic_Benefit_Status[1] == 1:
    outData.Withdrawal_Amount.append(outData.Maximum_Annual_Withdrawal[1])
else:
    outData.Withdrawal_Amount.append(0)

# L19 is AD19

# M19=MAX(0,I19-L19)
outData.AV_Post_Withdrawal.append(max(0, outData.AV_Pre_Withdrawal[1] - outData.Withdrawal_Amount[1]))

# P19=Rate.RiderCharge*M19
outData.Rider_Charge.append(inData.Rider_Charge * outData.AV_Post_Withdrawal[1])

# Q19=M19-P19
outData.AV_Post_Charges.append(outData.AV_Post_Withdrawal[1] - outData.Rider_Charge[1])

# To update U19 with an iterator
def iterative_solver(i):
    convergency = 10000
    generation = 50
    for item in range(0, generation):
        withdraw_base_data1 = outData.AV_Post_Death_Claims[i] if outData.Growth_Phase[i] == 1 else 0
        withdraw_base_data2 = outData.Withdrawal_Base[i - 1] * (1 - outData.qx[i]) + outData.Contribution[i]
        withdraw_base_data3 = outData.Withdrawal_Base[i - 1] * (1 - outData.qx[i]) * (1 + inData.Step_Up) + outData.Contribution[i] - \
                outData.MandE_Fund_Fees[i] - outData.Rider_Charge[i] if outData.Eligible_Step_Up[i] == 1 else 0
        # to calculate U19 first, then AC19
        outData.AV_Post_Death_Claims[i] = max(0, outData.AV_Post_Charges[i] - outData.Death_Payment[i])
        # AC19
        outData.Withdrawal_Base[i] = max(withdraw_base_data1, withdraw_base_data2, withdraw_base_data3)

        # AF19=AG19*AC19
        outData.Maximum_Annual_Withdrawal[i] = (outData.Maximum_Annual_Withdrawal_Rate[i] * outData.Withdrawal_Base[i])

        # AD19=IF(AK19=1,Rate.WD*AC19,IF(AL19=1,AF19,0))
        if outData.Withdrawal_Phase[i] == 1:
            outData.Withdrawal_Amount[i] = inData.Withdrawal_Rate * outData.Withdrawal_Base[i]
        elif outData.Automatic_Periodic_Benefit_Status[i] == 1:
            outData.Withdrawal_Amount[i] = outData.Maximum_Annual_Withdrawal[i]
        else:
            outData.Withdrawal_Amount[i] = 0

        #  L is AD

        # M19=MAX(0,I19-L19)
        outData.AV_Post_Withdrawal[i] = max(0, outData.AV_Pre_Withdrawal[i] - outData.Withdrawal_Amount[i])

        # P19=Rate.RiderCharge*M19
        outData.Rider_Charge[i] = inData.Rider_Charge * outData.AV_Post_Withdrawal[i]

        # Q19=M19-P19
        outData.AV_Post_Charges[i] = outData.AV_Post_Withdrawal[i] - outData.Rider_Charge[i]

        print("item：{0}  U ： {1}  AC： {2}  P：  {3}".format(item, outData.AV_Post_Death_Claims[i],
                                                           outData.Withdrawal_Base[i], outData.Rider_Charge[i]))


# to calculate the row19, 1 means row19
iterative_solver(1)
# after updating U and AC, we assign value to other variables

# to assign AB19，  AB19=MAX(0,AB18*(1-AT19)+D19-H19-L18-P19) for
# AB19=MAX(0,AB18*(1-AT19)+D19-H19-L18-P19)，在M19之后可求=============================
outData.Death_Benefit_base.append(max(0, outData.Death_Benefit_base[0] * (1 - outData.qx[1]) + outData.Contribution[1] -
                                      outData.MandE_Fund_Fees[1] - outData.Withdrawal_Amount[0] - outData.Rider_Charge[
                                          1]))

# J19=IF($I19=0,0,F19*($I19/$E19))
outData.Fund1_Pre_Withdrawal.append(0 if outData.AV_Pre_Withdrawal[1] == 0 else outData.Fund1_Pre_Fee[1] * (
        outData.AV_Pre_Withdrawal[1] / outData.AV_Pre_Fee[1]))

# K19=IF($I19=0,0,G19*($I19/$E19))
outData.Fund2_Pre_Withdrawal.append(0 if outData.AV_Pre_Withdrawal[1] == 0 else outData.Fund2_Pre_Fee[1] * (
        outData.AV_Pre_Withdrawal[1] / outData.AV_Pre_Fee[1]))

# N19=IF($M19=0,0,J19*($M19/$I19))
outData.Fund1_Post_Withdrawal.append(0 if outData.AV_Post_Withdrawal[1] == 0 else outData.Fund1_Pre_Withdrawal[1] * (
        outData.AV_Post_Withdrawal[1] / outData.AV_Pre_Withdrawal[1]))

#  O19==IF($M19=0,0,K19*($M19/$I19))
outData.Fund2_Post_Withdrawal.append(0 if outData.AV_Post_Withdrawal[1] == 0 else outData.Fund2_Pre_Withdrawal[1] * (
        outData.AV_Post_Withdrawal[1] / outData.AV_Pre_Withdrawal[1]))

# R19=IF($Q19=0,0,N19*($Q19/$M19))
outData.Fund1_Post_Charges.append(0 if outData.AV_Post_Charges[1] == 0 else outData.Fund1_Post_Withdrawal[1] * (
        outData.AV_Post_Charges[1] / outData.AV_Post_Withdrawal[1]))

# S19==IF($Q19=0,0,O19*($Q19/$M19))
outData.Fund2_Post_Charges.append(0 if outData.AV_Post_Charges[1] == 0 else outData.Fund2_Post_Withdrawal[1] * (
        outData.AV_Post_Charges[1] / outData.AV_Post_Withdrawal[1]))

# V19=IF($U19=0,0,R19*($U19/$Q19))
outData.Fund1_Post_Death_Claims.append(0 if outData.AV_Post_Death_Claims[1] == 0 else outData.Fund1_Post_Charges[1] * (
        outData.AV_Post_Death_Claims[1] / outData.AV_Post_Charges[1]))

#  W19=IF($U19=0,0,S19*($U19/$Q19))
outData.Fund2_Post_Death_Claims.append(0 if outData.AV_Post_Death_Claims[1] == 0 else outData.Fund2_Post_Charges[1] * (
        outData.AV_Post_Death_Claims[1] / outData.AV_Post_Charges[1]))

#  AQ19=AK19+AL19
outData.Rebalance_Indicator.append(outData.Withdrawal_Phase[1] + outData.Automatic_Periodic_Benefit_Status[1])

#  X19=IF(AQ19=1,U19*Fund.Reb.Target,V19)
outData.Fund1_Post_Rebalance.append(
    outData.AV_Post_Death_Claims[1] * inData.Fixed_Allocation_Funds_Automatic_Rebalancing_Target if
    outData.Rebalance_Indicator[1] == 1 else outData.Fund1_Post_Death_Claims[1])

#  Y19=Q19-X19
outData.Fund2_Post_Rebalance.append(outData.AV_Post_Charges[1] - outData.Fund1_Post_Rebalance[1])

# Z19=Z18*(1-AT19)
outData.ROP_Death_Base.append(outData.ROP_Death_Base[0] * (1 - outData.qx[1]))

# All columns before Z done =======================================================================================
#  AA19=MAX(0,T19-Q19)
outData.NAR_Death_Claims.append(max(0, outData.Death_Payment[1] - outData.AV_Post_Charges[1]))

# AE19=SUM(AD$18:AD19)
outData.Cumulative_Withdrawal.append(sum(outData.Withdrawal_Amount[:2]))

#  AU19=AA19
outData.Death_Claims = outData.NAR_Death_Claims

# AV19=MAX(AD19-U18,0)
outData.Withdrawal_Claims.append(max(0, outData.Withdrawal_Amount[1] - outData.AV_Post_Death_Claims[0]))

#  AW19=P19
outData.Rider_Charges = outData.Rider_Charge

# =========================Part Four: to calculate with Iterator ===================================================
for item in range(2, max_index):  # 从20行开始 # Start from row20
    #  F_item=X_item-1 * (1+AO_item)
    outData.Fund1_Pre_Fee.append(outData.Fund1_Post_Rebalance[item - 1] * (1 + outData.Fund1_Return[item]))
    # G_item=Y_item-1 * (1+AP_item)
    outData.Fund2_Pre_Fee.append(outData.Fund2_Post_Rebalance[item - 1] * (1 + outData.Fund2_Return[item]))
    # E_item=F_item+G_item
    outData.AV_Pre_Fee.append(outData.Fund1_Pre_Fee[item] + outData.Fund2_Pre_Fee[item])
    #  H_item=U_item-1*(Rate.MandE+Rate.FundFee)
    outData.MandE_Fund_Fees.append(outData.AV_Post_Death_Claims[item - 1] * (inData.MandE + inData.Fund_Fees))
    #  I_item=MAX(0,E_item+D_item-H_item)
    outData.AV_Pre_Withdrawal.append(
        max(0, outData.AV_Pre_Fee[item] + outData.Contribution[item] - outData.MandE_Fund_Fees[item]))

    #  AK_item=IF(AND(OR(C_item>Age.FirstWD,C_item>Age.AnnuityComm),U_item-1>0,C_item<Age.Death),1,0)
    if (outData.Age[item] > inData.First_Withdrawal_Age or outData.Age[item] > inData.Annuity_Commencement_Date) and \
            outData.AV_Post_Death_Claims[item - 1] > 0 and outData.Age[item] < inData.Last_Death_Age:
        outData.Withdrawal_Phase.append(1)
    else:
        outData.Withdrawal_Phase.append(0)

    #  AL_item=IF(C_item>=Age.Death,0,IF(AND(AK_item-1=1,U_item-1=0),1,AL_item-1))
    if outData.Age[item] >= inData.Last_Death_Age:
        outData.Automatic_Periodic_Benefit_Status.append(0)
    elif outData.Withdrawal_Phase[item - 1] == 1 and outData.AV_Post_Death_Claims[item - 1] == 0:
        outData.Automatic_Periodic_Benefit_Status.append(1)
    else:
        outData.Automatic_Periodic_Benefit_Status.append(outData.Automatic_Periodic_Benefit_Status[item - 1])

    #  AG_item=IF(AJ_item=1,0,IF(C_item>MAW.Age4,MAW.Rate4,IF(C_item>MAW.Age3,MAW.Rate3,IF(C_item>MAW.Age2,MAW.Rate2,IF(C_item>MAW.Age1,MAW.Rate1,0)))))
    if outData.Growth_Phase[item] == 1:
        outData.Maximum_Annual_Withdrawal_Rate.append(0)
    elif outData.Age[item] > inData.MAW_Age4:
        outData.Maximum_Annual_Withdrawal_Rate.append(inData.MAW_Rate4)
    elif outData.Age[item] > inData.MAW_Age3:
        outData.Maximum_Annual_Withdrawal_Rate.append(inData.MAW_Rate3)
    elif outData.Age[item] > inData.MAW_Age2:
        outData.Maximum_Annual_Withdrawal_Rate.append(inData.MAW_Rate2)
    elif outData.Age[item] > inData.MAW_Age1:
        outData.Maximum_Annual_Withdrawal_Rate.append(inData.MAW_Rate1)
    else:
        outData.Maximum_Annual_Withdrawal_Rate.append(0)

    #  T_item=IF(SUM(AJ_item:AM_item)=0,0,MAX(AB_item-1,Z_item-1)*AT_item)
    if outData.Growth_Phase[item] + outData.Withdrawal_Phase[item] + outData.Automatic_Periodic_Benefit_Status[item] + \
            outData.Last_Death[item] == 0:
        outData.Death_Payment.append(0)
    else:
        outData.Death_Payment.append(
            max(outData.Death_Benefit_base[item - 1], outData.ROP_Death_Base[item - 1]) * outData.qx[item])
    # =======================================start to iterate,with U,P,AC ========================
    # Initialize U item
    outData.AV_Post_Death_Claims.append(outData.AV_Post_Death_Claims[item - 1])

    #  AC19=MAX(IF(AJ19=1,U19,0),AC18*(1-AT19)+D19,IF(AI19=1,AC18*(1-AT19)*(1+Rate.StepUp)+D19-H19-P19,0))
    outData.Withdrawal_Base.append(outData.Withdrawal_Base[item - 1])

    # AF_item=AG_item*AC_item
    outData.Maximum_Annual_Withdrawal.append(
        outData.Maximum_Annual_Withdrawal_Rate[item] * outData.Withdrawal_Base[item])

    # AD_item=IF(AK_item=1,Rate.WD*AC_item,IF(AL_item=1,AF_item,0))
    if outData.Withdrawal_Phase[item] == 1:
        outData.Withdrawal_Amount.append(inData.Withdrawal_Rate * outData.Withdrawal_Base[item])
    elif outData.Automatic_Periodic_Benefit_Status[item] == 1:
        outData.Withdrawal_Amount.append(outData.Maximum_Annual_Withdrawal[item])
    else:
        outData.Withdrawal_Amount.append(0)
    # to assign L_item, L is AD

    #  M_item=MAX(0,I_item-L_item)
    outData.AV_Post_Withdrawal.append(max(0, outData.AV_Pre_Withdrawal[item] - outData.Withdrawal_Amount[item]))

    # P_item=Rate.RiderCharge*M_item
    outData.Rider_Charge.append(inData.Rider_Charge * outData.AV_Post_Withdrawal[item])

    # Q_item=M_item-P_item
    outData.AV_Post_Charges.append(outData.AV_Post_Withdrawal[item] - outData.Rider_Charge[item])

    # To update value of U_item========================================================================================================
    # to update value of 18+item
    iterative_solver(item)
    # After updating U and AC, assign value to rest of variables

    # AB_item=MAX(0,AB_item-1*(1-AT_item)+D_item-H_item-L_item-1-P_item)，calculate after M_item=============================
    outData.Death_Benefit_base.append(
        max(0, outData.Death_Benefit_base[item - 1] * (1 - outData.qx[item]) + outData.Contribution[item] -
            outData.MandE_Fund_Fees[item] - outData.Withdrawal_Amount[item - 1] -
            outData.Rider_Charge[item]))

    # J_item=IF($I_item=0,0,F_item*($I_item/$E_item))
    outData.Fund1_Pre_Withdrawal.append(0 if outData.AV_Pre_Withdrawal[item] == 0 else outData.Fund1_Pre_Fee[item] * (
            outData.AV_Pre_Withdrawal[item] / outData.AV_Pre_Fee[item]))

    # K_item=IF($I_item=0,0,G_item*($I_item/$E_item))
    outData.Fund2_Pre_Withdrawal.append(0 if outData.AV_Pre_Withdrawal[item] == 0 else outData.Fund2_Pre_Fee[item] * (
            outData.AV_Pre_Withdrawal[item] / outData.AV_Pre_Fee[item]))

    #  N_item=IF($M_item=0,0,J_item*($M_item/$I_item))
    outData.Fund1_Post_Withdrawal.append(
        0 if outData.AV_Post_Withdrawal[item] == 0 else outData.Fund1_Pre_Withdrawal[item] * (
                outData.AV_Post_Withdrawal[item] / outData.AV_Pre_Withdrawal[item]))

    # O_item==IF($M_item=0,0,K_item*($M_item/$I_item))
    outData.Fund2_Post_Withdrawal.append(
        0 if outData.AV_Post_Withdrawal[item] == 0 else outData.Fund2_Pre_Withdrawal[item] * (
                outData.AV_Post_Withdrawal[item] / outData.AV_Pre_Withdrawal[item]))

    #  R_item=IF($Q_item=0,0,N_item*($Q_item/$M_item))
    outData.Fund1_Post_Charges.append(
        0 if outData.AV_Post_Charges[item] == 0 else outData.Fund1_Post_Withdrawal[item] * (
                outData.AV_Post_Charges[item] / outData.AV_Post_Withdrawal[item]))

    #  S_item==IF($Q_item=0,0,O_item*($Q_item/$M_item))
    outData.Fund2_Post_Charges.append(
        0 if outData.AV_Post_Charges[item] == 0 else outData.Fund2_Post_Withdrawal[item] * (
                outData.AV_Post_Charges[item] / outData.AV_Post_Withdrawal[item]))

    #  V_item=IF($U_item=0,0,R_item*($U_item/$Q_item))
    outData.Fund1_Post_Death_Claims.append(
        0 if outData.AV_Post_Death_Claims[item] == 0 else outData.Fund1_Post_Charges[item] * (
                outData.AV_Post_Death_Claims[item] / outData.AV_Post_Charges[item]))

    #  W_item=IF($U_item=0,0,S_item*($U_item/$Q_item))
    outData.Fund2_Post_Death_Claims.append(
        0 if outData.AV_Post_Death_Claims[item] == 0 else outData.Fund2_Post_Charges[item] * (
                outData.AV_Post_Death_Claims[item] / outData.AV_Post_Charges[item]))

    #  AQ_item=AK_item+AL_item
    outData.Rebalance_Indicator.append(outData.Withdrawal_Phase[item] + outData.Automatic_Periodic_Benefit_Status[item])

    #  X_item=IF(AQ_item=1,U_item*Fund.Reb.Target,V_item)
    outData.Fund1_Post_Rebalance.append(
        outData.AV_Post_Death_Claims[item] * inData.Fixed_Allocation_Funds_Automatic_Rebalancing_Target if
        outData.Rebalance_Indicator[item] == 1 else outData.Fund1_Post_Death_Claims[item])

    #  Y_item=Q_item-X_item
    outData.Fund2_Post_Rebalance.append(outData.AV_Post_Charges[item] - outData.Fund1_Post_Rebalance[item])

    #  Z_item=Z_item-1*(1-AT_item)
    outData.ROP_Death_Base.append(outData.ROP_Death_Base[item - 1] * (1 - outData.qx[item]))

    # all items before Z is done  =======================================================================================
    # AA_item=MAX(0,T_item-Q_item)
    outData.NAR_Death_Claims.append(max(0, outData.Death_Payment[item] - outData.AV_Post_Charges[item]))

    # AE_item=SUM(AD$18:AD_item)
    outData.Cumulative_Withdrawal.append(sum(outData.Withdrawal_Amount[:item + 1]))

    # AU_item=AA_item
    outData.Death_Claims = outData.NAR_Death_Claims

    #  AV_item=MAX(AD_item-U_item-1,0)
    outData.Withdrawal_Claims.append(max(0, outData.Withdrawal_Amount[item] - outData.AV_Post_Death_Claims[item - 1]))

    #  AW_item=P_item
    outData.Rider_Charges = outData.Rider_Charge

    # To calculate PV_DB_Claim, AU16=DF*outData.Death_Claims
    outData.PV_DB_Claim += outData.DF[item] * outData.Death_Claims[item]
    # To calculate PV_WB_Claim, AV16== DF*outData.Withdrawal_Claims
    outData.PV_WB_Claim += outData.DF[item] * outData.Withdrawal_Claims[item]
    # To calculate PV_RC, AW16== DF*outData.Rider_Charges
    outData.PV_RC += outData.DF[item] * outData.Rider_Charges[item]

# ===============================================写出到excel中===================================================
# Workbook() takes one, non-optional, argument
# which is the filename that we want to create.
workbook = xlsxwriter.Workbook('CashFlowOut.xlsx')
worksheet = workbook.add_worksheet("cashFlowOutput")
workbook.close()

# print output back to excel sheet
wb_CashflowOut = vb.load_workbook('CashFlowOut.xlsx')
wsOutput = wb_CashflowOut["cashFlowOutput"]
wsOutput.cell(1, 1).value = "Year"
wsOutput.cell(1, 2).value = "Anniversary"
wsOutput.cell(1, 3).value = "Age"
wsOutput.cell(1, 4).value = "Contribution"
wsOutput.cell(1, 5).value = "AV Pre-Fee"
wsOutput.cell(1, 6).value = "Fund1 Pre-Fee"
wsOutput.cell(1, 7).value = "Fund2 Pre-Fee"
wsOutput.cell(1, 8).value = "M&E/Fund Fees"
wsOutput.cell(1, 9).value = "AV Pre-Withdrawal"
wsOutput.cell(1, 10).value = "Fund1 Pre-Withdrawal"
wsOutput.cell(1, 11).value = "Fund2 Pre-Withdrawal"
wsOutput.cell(1, 12).value = "Withdrawal Amount"
wsOutput.cell(1, 13).value = "AV Post-Withdrawal"
wsOutput.cell(1, 14).value = "Fund1 Post-Withdrawal"
wsOutput.cell(1, 15).value = "Fund2 Post-Withdrawal"
wsOutput.cell(1, 16).value = "Rider Charge"
wsOutput.cell(1, 17).value = "AV Post-Charges"
wsOutput.cell(1, 18).value = "Fund1 Post-Charges"
wsOutput.cell(1, 19).value = "Fund2 Post-Charges"
wsOutput.cell(1, 20).value = "Death Payments"
wsOutput.cell(1, 21).value = "AV Post-Death Claims"
wsOutput.cell(1, 22).value = "Fund1 Post-Death Claims"
wsOutput.cell(1, 23).value = "Fund2 Post-Death Claims"
wsOutput.cell(1, 24).value = "Fund1 Post-Rebalance"
wsOutput.cell(1, 25).value = "Fund2 Post-Rebalance"
wsOutput.cell(1, 26).value = "ROP Death Base"

wsOutput.cell(1, 27).value = "NAR Death Claims"
wsOutput.cell(1, 28).value = "Death Benefit base"
wsOutput.cell(1, 29).value = "Withdrawal Base"
wsOutput.cell(1, 30).value = "Withdrawal Amount"
wsOutput.cell(1, 31).value = "Cumulative Withdrawal"
wsOutput.cell(1, 32).value = "Maximum Annual Withdrawal"
wsOutput.cell(1, 33).value = "Maximum Annual Withdrawal Rate"

wsOutput.cell(1, 35).value = "Eligible_Step_UP"
wsOutput.cell(1, 36).value = "Growth_Phase"
wsOutput.cell(1, 37).value = "WithDrawal_Phase"
wsOutput.cell(1, 38).value = "Automatic Periodic Benefit Status"
wsOutput.cell(1, 39).value = "Last_Death"

wsOutput.cell(1, 41).value = "Fund1 Return"
wsOutput.cell(1, 42).value = "Fund2 Return"
wsOutput.cell(1, 43).value = "Rebalance indicator"
wsOutput.cell(1, 44).value = "DF"

wsOutput.cell(1, 46).value = "qx"
wsOutput.cell(1, 47).value = "Death Claims"
wsOutput.cell(1, 48).value = "Withdrawal Claims"
wsOutput.cell(1, 49).value = "Rider Charges"

# total DB,WB and RC
wsOutput.cell(1, 50).value = "PV_DB_Claim"
wsOutput.cell(1, 51).value = "PV_WB_Claim"
wsOutput.cell(1, 52).value = "PV_RC"

wsOutput.cell(2, 50).value = outData.PV_DB_Claim
wsOutput.cell(2, 51).value = outData.PV_WB_Claim
wsOutput.cell(2, 52).value = outData.PV_RC

for item in range(0, max_index):
    # A、B、C、D
    wsOutput.cell(item + 2, 1).value = outData.Year[item]
    wsOutput.cell(item + 2, 2).value = outData.Anniversary[item]
    wsOutput.cell(item + 2, 3).value = outData.Age[item]
    wsOutput.cell(item + 2, 4).value = outData.Contribution[item]

    # E、F、G、H、I、G
    wsOutput.cell(item + 2, 5).value = outData.AV_Pre_Fee[item]
    wsOutput.cell(item + 2, 6).value = outData.Fund1_Pre_Fee[item]
    wsOutput.cell(item + 2, 7).value = outData.Fund2_Pre_Fee[item]
    wsOutput.cell(item + 2, 8).value = outData.MandE_Fund_Fees[item]
    wsOutput.cell(item + 2, 9).value = outData.AV_Pre_Withdrawal[item]
    wsOutput.cell(item + 2, 10).value = outData.Fund1_Pre_Withdrawal[item]
    # K、L、M、N、O
    wsOutput.cell(item + 2, 11).value = outData.Fund2_Pre_Withdrawal[item]
    wsOutput.cell(item + 2, 12).value = outData.Withdrawal_Amount[item]
    wsOutput.cell(item + 2, 13).value = outData.AV_Post_Withdrawal[item]
    wsOutput.cell(item + 2, 14).value = outData.Fund1_Post_Withdrawal[item]
    wsOutput.cell(item + 2, 15).value = outData.Fund2_Post_Withdrawal[item]
    # P、Q、R、S、T、
    wsOutput.cell(item + 2, 16).value = outData.Rider_Charge[item]
    wsOutput.cell(item + 2, 17).value = outData.AV_Post_Charges[item]
    wsOutput.cell(item + 2, 18).value = outData.Fund1_Post_Charges[item]
    wsOutput.cell(item + 2, 19).value = outData.Fund2_Post_Charges[item]
    wsOutput.cell(item + 2, 20).value = outData.Death_Payment[item]
    # U、V、W、X、Y、Z
    wsOutput.cell(item + 2, 21).value = outData.AV_Post_Death_Claims[item]
    wsOutput.cell(item + 2, 22).value = outData.Fund1_Post_Death_Claims[item]
    wsOutput.cell(item + 2, 23).value = outData.Fund2_Post_Death_Claims[item]
    wsOutput.cell(item + 2, 24).value = outData.Fund1_Post_Rebalance[item]
    wsOutput.cell(item + 2, 25).value = outData.Fund2_Post_Rebalance[item]
    wsOutput.cell(item + 2, 26).value = outData.ROP_Death_Base[item]

    # AA、AB、AC、AD、AE、AF、AG
    wsOutput.cell(item + 2, 27).value = outData.NAR_Death_Claims[item]
    wsOutput.cell(item + 2, 28).value = outData.Death_Benefit_base[item]
    wsOutput.cell(item + 2, 29).value = outData.Withdrawal_Base[item]
    wsOutput.cell(item + 2, 30).value = outData.Withdrawal_Amount[item]
    wsOutput.cell(item + 2, 31).value = outData.Cumulative_Withdrawal[item]
    wsOutput.cell(item + 2, 32).value = outData.Maximum_Annual_Withdrawal[item]
    wsOutput.cell(item + 2, 33).value = outData.Maximum_Annual_Withdrawal_Rate[item]

    # AI、AJ、AK、AL、AM
    wsOutput.cell(item + 2, 35).value = outData.Eligible_Step_Up[item]
    wsOutput.cell(item + 2, 36).value = outData.Growth_Phase[item]
    wsOutput.cell(item + 2, 37).value = outData.Withdrawal_Phase[item]
    wsOutput.cell(item + 2, 38).value = outData.Automatic_Periodic_Benefit_Status[item]

    wsOutput.cell(item + 2, 39).value = outData.Last_Death[item]

    # AO AP AQ Fund1 Return Fund2 Return Rebalance Indicator
    wsOutput.cell(item + 2, max_index).value = outData.Fund1_Return[item]
    wsOutput.cell(item + 2, max_index).number_format = '0.00%'
    # AP
    wsOutput.cell(item + 2, 42).value = outData.Fund2_Return[item]
    wsOutput.cell(item + 2, 42).number_format = '0.00%'
    # AQ
    wsOutput.cell(item + 2, 43).value = outData.Rebalance_Indicator[item]
    # AR
    wsOutput.cell(item + 2, 44).value = outData.DF[item]

    # AT、AU、AV、AW
    wsOutput.cell(item + 2, 46).value = outData.qx[item]
    wsOutput.cell(item + 2, 47).value = outData.Death_Claims[item]
    wsOutput.cell(item + 2, 48).value = outData.Withdrawal_Claims[item]
    wsOutput.cell(item + 2, 49).value = outData.Rider_Charges[item]

wb_CashflowOut.save('CashflowOut.xlsx')
wb_CashflowOut.close()
